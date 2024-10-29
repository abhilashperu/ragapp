from io import BytesIO
import streamlit as st
#from dotenv import load_dotenv
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.embeddings import OpenAIEmbeddings, OllamaEmbeddings, HuggingFaceInstructEmbeddings
from langchain_community.vectorstores import FAISS
from langchain_community.chat_models import ChatOpenAI, ChatOllama
from pydantic.v1 import BaseModel, ConfigDict
from langchain.memory import ConversationBufferMemory
from langchain.chains.conversational_retrieval.base import ConversationalRetrievalChain
from sentence_transformers import SentenceTransformer
from htmlTemplates import css, bot_template, user_template
import pickle
#from langchain_community.llms import huggingface_hub
from langchain_community.llms.huggingface_hub import HuggingFaceHub
#from huggingface_hub import hf_hub_download
from InstructorEmbedding import INSTRUCTOR
from sentence_transformers import SentenceTransformer
from langchain.chains import LLMChain
from langchain_community.llms import huggingface_endpoint
from transformers import AutoTokenizer, AutoModelForCausalLM
from langchain.prompts import ChatPromptTemplate, PromptTemplate
import warnings
import os
import numpy as np
from pydantic.v1 import root_validator, validator
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import pandas as pd
from openpyxl import load_workbook

# Load environment variables
#load_dotenv()
os.environ['CURL_CA_BUNDLE'] = ''#'C:\\Users\\320124527\\Downloads\\huggingface.co.crt'
os.environ['REQUESTS_CA_BUNDLE'] = ''
#sec_key = "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" 
sec_key = os.environ['HUGGINGFACEHUB_API_TOKEN']

# Suppress warnings
warnings.filterwarnings("ignore")

# Initialize SentenceTransformer model
model = SentenceTransformer('paraphrase-MiniLM-L6-v2')

def get_pdf(pdf_docs):
    raw_text = ""
    for pdf in pdf_docs:
        pdf_reader = PdfReader(pdf)
        store_name = pdf.name[:-4]
        for page in pdf_reader.pages:
            raw_text += page.extract_text()
    return raw_text

def get_text_chunks(raw_text):
    text_splitter = CharacterTextSplitter(separator="\n",
                                          chunk_size = 2000,
                                          chunk_overlap = 250,
                                          length_function=len)
    chunks = text_splitter.split_text(raw_text)
    return chunks

def get_vector_store(text_chunks):
    #model = SentenceTransformer("hkunlp/instructor-xl")
    #embeddings = model.encode(text_chunks)
    #np.save('embeddings.npy', embeddings)
    #embeddings.save('local_embedding_model')
    #vectorstore = FAISS.from_embeddings(texts = text_chunks, embedding = embeddings)
    model_name = "hkunlp/instructor-large"
    model_kwargs = {'device': 'cpu'}
    encode_kwargs = {'normalize_embeddings': True}
    hf = HuggingFaceInstructEmbeddings(
        model_name=model_name,
        model_kwargs=model_kwargs,
        encode_kwargs=encode_kwargs)
    vectorstore = FAISS.from_texts(texts = text_chunks, embedding = hf)
    return vectorstore
    print(vectorstore)


def get_conversation_chain(vectorstore):

    #tokenizer = AutoTokenizer.from_pretrained("mistralai/Mistral-7B-Instruct-v0.2")
    #model = AutoModelForCausalLM.from_pretrained("mistralai/Mistral-7B-Instruct-v0.2")
    #tokenizer.save_pretrained("local_llm_model")
    #model.save_pretrained("local_llm_model")

    repo_id="https://api-inference.huggingface.co/models/mistralai/Mistral-7B-Instruct-v0.2"
    #repo_id = "https://api-inference.huggingface.co/models/google/flan-t5-xxl"
    #llm = huggingface_hub(repo_id=repo_id, huggingfacehub_api_token=sec_key)
    #llm = HuggingFaceHub(repo_id = repo_id)
    llm = huggingface_endpoint.HuggingFaceEndpoint(
    endpoint_url=repo_id,
    max_length=1000,
    temperature=0.5,
    huggingfacehub_api_token=sec_key
    )
    #llm = hf_hub_download(repo_id=repo_id, filename="config.json"), model_kwargs={"temperature":0.5, "max_length":10024, "max_new_tokens":512}

    template = """Given the following conversation respond to the best of your ability in a pirate voice and end every sentence with Ay Ay Matey
     Additionally, if the conversation requests a table or tabular form, provide it in the following format where Protocol ID refers to sequence number, 
     Name is Parameter Name, Description referes to information about that parameter, Unit refers to Unit of the Parameter, Precision is how many decimal points of that parameter, 
     Minimum value refers to minimum range value of that parameter, Maximum value refers to maximum range value of that parameter, ID is a number in multiple of 10 starting from 10, 
     Scoped is always 'Y' and Devices is always '*'. If any data is not available then output it as 'N/A':
     make it into a table with below columns and output it
   
    | | | | Protocol    |
    |ID | Scoped | Devices | Protocol ID | Name | Description | Unit | Precision | Minimum Value | Maximum Value |
    |10 | Y | * | protocol ID 1 | parameter_name_1 | parameter_description | unit_1 | Precision_1 | Minimum_value_1 | Maximum_value_1 |
    |20 | Y | * | protocol ID 2 | parameter_name_2 | parameter_description | unit_2 | Precision_2 | Minimum_value_2 | Maximum_value_2 |
    
    
    Chat History:
    {chat_history}
    Follow Up Input: {question}
    Standalone question:"""

    PROMPT = PromptTemplate(
    input_variables=["chat_history", "question"], 
    template=template
    )


    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain

def calculate_similarity(query, response):
    vectorizer = TfidfVectorizer().fit_transform([query, response])
    vectors = vectorizer.toarray()
    cosine_sim = cosine_similarity(vectors)
    return cosine_sim[0, 1]

def semantic_similarity(query, response):
    embeddings = model.encode([query, response])
    return cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]


def extract_protocol_data(content):
    protocol_data = []
    lines = content.strip().split("\n")

    # Detect header and start of the data
    header_found = False
    headers = []

    for line in lines:
        if '|' in line:
            line_content = [col.strip() for col in line.split('|') if col.strip()]
            if len(line_content) > 1:  # We found a potential header or data line
                if not header_found:
                    headers = line_content
                    header_found = True
                else:
                    if set(line.strip()) != {'-'}:  # Exclude separator lines like "-----"
                        # Ensure the number of data columns matches the number of headers
                        if len(line_content) == len(headers):
                            protocol_data.append(dict(zip(headers, line_content)))
    
    return protocol_data

def fill_template(protocol_data):
    wb = load_workbook('template.xlsx')
    ws = wb.active
    
    start_row_idx = 3
    start_col_idx = ws['D' + str(start_row_idx)].col_idx

    protocol_data = protocol_data[2:]  # Skip the first 2 rows of the output
    
    for idx, protocol in enumerate(protocol_data):
        for col_num, header in enumerate(protocol.keys(), start=start_col_idx):
            cell = ws.cell(row=start_row_idx + idx, column=col_num)
            cell.value = protocol.get(header, 'N/A')

    return wb


def handle_userinput(user_question):
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']
    protocol_data = []
    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            cosine_similarity_score = calculate_similarity(user_question, message.content)
            semantic_similarity_score = semantic_similarity(user_question, message.content)
            formatted_bot_template = (bot_template.replace("{{MSG}}", message.content)  .replace("{{COSINE_SCORE}}", f"{cosine_similarity_score:.2f}").replace("{{SEMANTIC_SCORE}}", f"{semantic_similarity_score:.2f}"))
            st.write(formatted_bot_template, unsafe_allow_html=True)
            # Extract protocol data from the bot response content
            protocol_data = extract_protocol_data(message.content)
            st.session_state.protocol_data = protocol_data

            if protocol_data:
                st.subheader("Extracted Protocol Data from Response")
                st.table(protocol_data)


def main():
    #load_dotenv()
    st.set_page_config(page_title="Get Answers from your Documents", page_icon=":books:")
    st.write(css, unsafe_allow_html=True)

    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None
    if "protocol_data" not in st.session_state:
        st.session_state.protocol_data = None   

    st.header("Chat with your Protocols :books:")
    

    with st.sidebar:
        st.subheader("Your Protocols")
        pdf_docs = st.file_uploader("Upload your PDFs here and click on 'Process'", accept_multiple_files=True)
        #template_file = st.file_uploader("Upload your Excel Template", type=["xlsx"])
        if st.button("Process"):
            with st.spinner("Processing"):
                # get the PDF
                raw_text = get_pdf(pdf_docs)

                # get the text chunks
                text_chunks = get_text_chunks(raw_text)
                
                # create the vector store
                vector_store = get_vector_store(text_chunks)

                # create conversation chain
                st.session_state.conversation = get_conversation_chain(
                    vector_store)
            
                        
    user_question = st.text_input("Ask a question about your Protocols:")
    if user_question:
       handle_userinput(user_question)

     # Export to XLSX button for extracted protocol data
    if st.session_state.protocol_data:
        if st.button("Export Extracted Data to XLSX"):
            wb = fill_template(st.session_state.protocol_data)
            
            save_path = "updated_template.xlsx"
            wb.save(save_path)

            with open(save_path, "rb") as f:
                st.download_button(
                    label="Download Extracted Data XLSX",
                    data=f,
                    file_name='extracted_protocol_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    
if __name__ == '__main__':
    main()